from fastapi import FastAPI, UploadFile, File, Form, HTTPException, Body
from fastapi.middleware.cors import CORSMiddleware
from fastapi.staticfiles import StaticFiles
from fastapi.responses import RedirectResponse
from typing import List, Optional, Dict, Any, Sequence, Tuple
import pandas as pd
from io import BytesIO
import uvicorn
import os
import json
from datetime import datetime
from transformations import TransformationEngine
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill
from openpyxl.formatting.rule import FormulaRule
from openpyxl.chart import LineChart, Reference
from openpyxl.formatting.rule import ColorScaleRule, DataBarRule, IconSetRule
import hashlib
import platform
import sys
import time
try:
    import duckdb  # optional
except Exception:
    duckdb = None
try:
    import redis as redis_lib
except Exception:
    redis_lib = None
try:
    import httpx  # optional
except Exception:
    httpx = None

app = FastAPI(title="Universal Data Uploader", version="1.0.0")

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

_static_dir = os.path.join(os.path.dirname(__file__), "static")
if os.path.isdir(_static_dir):
    app.mount("/app", StaticFiles(directory=_static_dir, html=True), name="app")

def _read_ai_file_config() -> dict:
    """
    Load optional OpenAI configuration from ai_config.json in the project root.
    Returns an empty dict if the file is missing or invalid.
    File values take precedence over environment variables.
    """
    path = os.path.join(os.path.dirname(__file__), "ai_config.json")
    try:
        if os.path.exists(path):
            with open(path, "r", encoding="utf-8") as f:
                data = json.load(f)
                return data if isinstance(data, dict) else {}
    except Exception:
        pass
    return {}

_file_cfg = _read_ai_file_config()
_env_enabled = os.getenv("AI_SUGGESTIONS_ENABLED", "1")
AI_ENABLED = bool(_file_cfg.get("AI_SUGGESTIONS_ENABLED", None)) if "AI_SUGGESTIONS_ENABLED" in _file_cfg else (_env_enabled not in ("0", "false", "False", ""))
OPENAI_API_KEY = _file_cfg.get("OPENAI_API_KEY") or os.getenv("OPENAI_API_KEY")
OPENAI_API_BASE = _file_cfg.get("OPENAI_API_BASE") or os.getenv("OPENAI_API_BASE", "https://api.openai.com/v1")
OPENAI_MODEL = _file_cfg.get("OPENAI_MODEL") or os.getenv("OPENAI_MODEL", "gpt-4o-mini")
GEMINI_API_KEY = _file_cfg.get("GEMINI_API_KEY") or os.getenv("GEMINI_API_KEY")
GEMINI_API_BASE = _file_cfg.get("GEMINI_API_BASE") or os.getenv("GEMINI_API_BASE", "https://generativelanguage.googleapis.com/v1beta")
GEMINI_MODEL = _file_cfg.get("GEMINI_MODEL") or os.getenv("GEMINI_MODEL", "gemini-1.5-flash")
CUSTOM_AI_URL = _file_cfg.get("CUSTOM_AI_URL") or os.getenv("CUSTOM_AI_URL")
CUSTOM_AI_KEY = _file_cfg.get("CUSTOM_AI_KEY") or os.getenv("CUSTOM_AI_KEY")
CUSTOM_AI_HEADER_NAME = _file_cfg.get("CUSTOM_AI_HEADER_NAME") or os.getenv("CUSTOM_AI_HEADER_NAME")
CUSTOM_AI_HEADER_VALUE = _file_cfg.get("CUSTOM_AI_HEADER_VALUE") or os.getenv("CUSTOM_AI_HEADER_VALUE")
CUSTOM_AI_MODEL = _file_cfg.get("CUSTOM_AI_MODEL") or os.getenv("CUSTOM_AI_MODEL", "balanced")
if isinstance(CUSTOM_AI_URL, str):
    CUSTOM_AI_URL = CUSTOM_AI_URL.strip()
if isinstance(CUSTOM_AI_KEY, str):
    CUSTOM_AI_KEY = CUSTOM_AI_KEY.strip()
if isinstance(CUSTOM_AI_HEADER_NAME, str):
    CUSTOM_AI_HEADER_NAME = CUSTOM_AI_HEADER_NAME.strip()
if isinstance(CUSTOM_AI_HEADER_VALUE, str):
    CUSTOM_AI_HEADER_VALUE = CUSTOM_AI_HEADER_VALUE.strip()

def _dedup_steps(steps):
    """
    Deduplicate pipeline steps by (type, config) while preserving order.
    Returns a new list of {type, config} entries.
    """
    seen = set()
    out = []
    for s in steps or []:
        t = s.get("type")
        cfg = s.get("config", {}) or {}
        try:
            key = json.dumps({"type": t, "config": cfg}, sort_keys=True)
        except Exception:
            key = f"{t}:{cfg}"
        if key in seen:
            continue
        seen.add(key)
        out.append({"type": t, "config": cfg})
    return out

def _ai_propose_recipe(columns: list, sample_rows: list) -> dict:
    """
    Ask an OpenAI-compatible endpoint for domain-specific transformation steps.
    Returns {"steps":[{type,config},...], "suggestions":[...]} or {} on failure.
    """
    if not AI_ENABLED or not httpx:
        return {}
    if CUSTOM_AI_URL:
        try:
            sys_prompt = (
                "You generate data transformation steps for a tabular dataset of any domain. "
                "Return only JSON with keys 'steps' and 'suggestions'. "
                "Allowed step types with configs: "
                "standardizeDate{col}, sortByDate{col,order}, calculateNetProfit{}, "
                "currencyToFloat{columns}, addColumnFormula{dest,expr}, "
                "groupByAggregate{by,aggs}, dropDuplicates{subset,keep}, "
                "detectOutliers{columns,k}, filterExpr{expr}, filterDebitCredit{col,include}, "
                "aggregatePLByPeriod{date_col,freq,revenue_col,cogs_col,expenses_col}, "
                "mergeColumns{cols,dest,sep}. "
                "Infer the dataset domain from column names. If finance columns exist, include P&L related steps; "
                "otherwise propose domain‑agnostic cleaning, validation, grouping and enrichment steps. "
                "Use column names exactly as provided."
            )
            headers = {"Content-Type": "application/json"}
            if CUSTOM_AI_HEADER_NAME and CUSTOM_AI_HEADER_VALUE:
                headers[CUSTOM_AI_HEADER_NAME] = CUSTOM_AI_HEADER_VALUE
            elif CUSTOM_AI_KEY:
                headers["Authorization"] = f"Bearer {CUSTOM_AI_KEY}"
            body = {
                "model": CUSTOM_AI_MODEL or "balanced",
                "messages": [
                    {"role": "system", "content": sys_prompt},
                    {"role": "user", "content": f"Columns:\n{json.dumps(columns)}\nSample:\n{json.dumps(sample_rows[:20] if sample_rows else [])}\nReturn JSON only."}
                ],
                "max_tokens": 1200,
                "temperature": 0.2
            }
            with httpx.Client(timeout=20.0) as client:
                resp = client.post(CUSTOM_AI_URL, headers=headers, json=body)
            resp.raise_for_status()
            rj = resp.json()
            parsed = {}
            if isinstance(rj, dict) and ("steps" in rj or "suggestions" in rj):
                parsed = rj
            else:
                content = None
                try:
                    content = rj.get("message", {}).get("content")
                except Exception:
                    content = None
                if not content:
                    try:
                        content = rj.get("choices", [{}])[0].get("message", {}).get("content")
                    except Exception:
                        content = None
                if not content:
                    try:
                        content = rj.get("content") or rj.get("reply") or rj.get("text")
                    except Exception:
                        content = None
                if not content:
                    try:
                        msgs = rj.get("messages")
                        if isinstance(msgs, list) and msgs:
                            last = None
                            for m in msgs[::-1]:
                                if isinstance(m, dict) and (m.get("role") in ("assistant", "system", "bot")) and m.get("content"):
                                    last = m.get("content"); break
                            if not last:
                                for m in msgs[::-1]:
                                    if isinstance(m, dict) and m.get("content"):
                                        last = m.get("content"); break
                            content = last
                    except Exception:
                        content = None
                if not content:
                    try:
                        content = rj.get("data", {}).get("content") or rj.get("result", {}).get("content") or rj.get("response")
                    except Exception:
                        content = None
                if content:
                    try:
                        txt = content
                        if isinstance(txt, str) and "```" in txt:
                            s = txt.find("```")
                            e = txt.rfind("```")
                            if e > s:
                                fence = txt[s+3:e]
                                if fence.strip().startswith("json"):
                                    fence = fence.strip()[4:]
                                txt = fence.strip()
                        parsed = json.loads(txt)
                    except Exception:
                        parsed = {}
            raw_steps = parsed.get("steps") or parsed.get("transforms") or []
            raw_suggestions = parsed.get("suggestions") or parsed.get("hints") or []
            steps = _dedup_steps(raw_steps)
            suggestions = list(dict.fromkeys(raw_suggestions))
            return {"steps": steps, "suggestions": suggestions, "provider": "custom"}
        except Exception:
            pass
    if GEMINI_API_KEY:
        try:
            sys_prompt = (
                "You generate data transformation steps for a tabular dataset of any domain. "
                "Return only JSON with keys 'steps' and 'suggestions'. "
                "Allowed step types with configs: "
                "standardizeDate{col}, sortByDate{col,order}, calculateNetProfit{}, "
                "currencyToFloat{columns}, addColumnFormula{dest,expr}, "
                "groupByAggregate{by,aggs}, dropDuplicates{subset,keep}, "
                "detectOutliers{columns,k}, filterExpr{expr}, filterDebitCredit{col,include}, "
                "aggregatePLByPeriod{date_col,freq,revenue_col,cogs_col,expenses_col}, "
                "mergeColumns{cols,dest,sep}. "
                "Infer the dataset domain from column names. If finance columns exist, include P&L related steps; "
                "otherwise propose domain‑agnostic cleaning, validation, grouping and enrichment steps. "
                "Use column names exactly as provided."
            )
            url = f"{GEMINI_API_BASE.rstrip('/')}/models/{GEMINI_MODEL}:generateContent?key={GEMINI_API_KEY}"
            body = {
                "contents": [
                    {
                        "role": "user",
                        "parts": [
                            {
                                "text": sys_prompt + "\n\n" +
                                "Columns:\n" + json.dumps(columns) + "\n" +
                                "Sample:\n" + json.dumps(sample_rows[:20] if sample_rows else []) + "\n" +
                                "Return JSON only."
                            }
                        ]
                    }
                ],
                "generationConfig": {"temperature": 0.2}
            }
            with httpx.Client(timeout=15.0) as client:
                resp = client.post(url, json=body)
            resp.raise_for_status()
            data = resp.json()
            text = ""
            try:
                text = data["candidates"][0]["content"]["parts"][0]["text"]
            except Exception:
                text = ""
            parsed = json.loads(text) if text else {}
            steps = _dedup_steps(parsed.get("steps", []))
            suggestions = list(dict.fromkeys(parsed.get("suggestions", [])))
            return {"steps": steps, "suggestions": suggestions, "provider": "gemini"}
        except Exception:
            return {}
    if not OPENAI_API_KEY:
        return {}
    try:
        sys_prompt = (
            "You generate data transformation steps for a tabular dataset of any domain. "
            "Return only JSON with keys 'steps' and 'suggestions'. "
            "Allowed step types with configs: "
            "standardizeDate{col}, sortByDate{col,order}, calculateNetProfit{}, "
            "currencyToFloat{columns}, addColumnFormula{dest,expr}, "
            "groupByAggregate{by,aggs}, dropDuplicates{subset,keep}, "
            "detectOutliers{columns,k}, filterExpr{expr}, filterDebitCredit{col,include}, "
            "aggregatePLByPeriod{date_col,freq,revenue_col,cogs_col,expenses_col}, "
            "mergeColumns{cols,dest,sep}. "
            "Infer the dataset domain from column names. If finance columns exist, include P&L related steps; "
            "otherwise propose domain‑agnostic cleaning, validation, grouping and enrichment steps. "
            "Use column names exactly as provided."
        )
        user_payload = {
            "columns": columns,
            "sample": sample_rows[:20] if sample_rows else [],
        }
        url = f"{OPENAI_API_BASE.rstrip('/')}/chat/completions"
        headers = {"Authorization": f"Bearer {OPENAI_API_KEY}", "Content-Type": "application/json"}
        body = {
            "model": OPENAI_MODEL,
            "messages": [
                {"role": "system", "content": sys_prompt},
                {"role": "user", "content": f"Columns:\n{json.dumps(columns)}\nSample:\n{json.dumps(user_payload['sample'])}\nReturn JSON only."}
            ],
            "temperature": 0.2,
        }
        with httpx.Client(timeout=15.0) as client:
            resp = client.post(url, headers=headers, json=body)
        resp.raise_for_status()
        data = resp.json()
        content = data["choices"][0]["message"]["content"]
        parsed = json.loads(content)
        steps = _dedup_steps(parsed.get("steps", []))
        suggestions = list(dict.fromkeys(parsed.get("suggestions", [])))
        return {"steps": steps, "suggestions": suggestions, "provider": "openai"}
    except Exception:
        return {}

class DataFactory:
    """
    Utilities to read tabular data from supported formats and Excel sheets.
    Provides helpers for full DataFrame reads and sheet-aware listing/preview.
    """
    @staticmethod
    def _to_bytes(data: UploadFile) -> bytes:
        """Return full file bytes from an UploadFile without altering the caller's pointer."""
        data.file.seek(0)
        return data.file.read()

    @staticmethod
    def detect_extension(filename: str) -> str:
        """Infer file extension type supported by the backend."""
        lower = filename.lower()
        if lower.endswith(".xlsx"):
            return "xlsx"
        if lower.endswith(".xls"):
            return "xls"
        if lower.endswith(".csv"):
            return "csv"
        if lower.endswith(".json"):
            return "json"
        if lower.endswith(".parquet"):
            return "parquet"
        raise ValueError("Unsupported file format")

    @staticmethod
    def list_sheets(file: UploadFile) -> List[str]:
        """List sheet names from an uploaded Excel workbook."""
        content = DataFactory._to_bytes(file)
        try:
            ext = DataFactory.detect_extension(file.filename)
            engine = "xlrd" if ext == "xls" else "openpyxl"
            xls = pd.ExcelFile(BytesIO(content), engine=engine)
            return xls.sheet_names
        except Exception as e:
            raise HTTPException(status_code=400, detail=str(e))

    @staticmethod
    def read_df(
        file: UploadFile,
        sheets: Optional[Sequence[str]] = None,
        merge_identical: bool = False,
    ) -> pd.DataFrame:
        """
        Read the full dataset into a DataFrame with optional Excel sheet selection/merging.
        """
        ext = DataFactory.detect_extension(file.filename)
        content = DataFactory._to_bytes(file)
        if ext in ("xlsx", "xls"):
            try:
                engine = "xlrd" if ext == "xls" else "openpyxl"
                xls = pd.ExcelFile(BytesIO(content), engine=engine)
            except Exception as e:
                raise HTTPException(status_code=400, detail=str(e))
            sheet_names = xls.sheet_names
            chosen = list(sheets) if sheets else sheet_names
            frames: Dict[str, pd.DataFrame] = {sn: pd.read_excel(xls, sheet_name=sn, engine=engine) for sn in chosen}
            if merge_identical and len(frames) > 1:
                groups: Dict[tuple, List[str]] = {}
                for name, df in frames.items():
                    key = tuple([c.strip().lower() for c in df.columns])
                    groups.setdefault(key, []).append(name)
                if groups:
                    largest_key = max(groups.keys(), key=lambda k: len(groups[k]))
                    merge_list = groups[largest_key]
                    return pd.concat([frames[n] for n in merge_list], ignore_index=True)
            # If schemas are same, concat; else return first sheet by default
            if len(frames) == 1:
                return next(iter(frames.values()))
            keys = [tuple([c.strip().lower() for c in df.columns]) for df in frames.values()]
            if all(k == keys[0] for k in keys):
                return pd.concat(list(frames.values()), ignore_index=True)
            # Fallback: first sheet
            first_name = chosen[0]
            return frames[first_name]
        if ext == "csv":
            data = content.replace(b"\x00", b"") if b"\x00" in content else content
            for kwargs in [
                {"low_memory": False},
                {"encoding": "utf-8-sig", "engine": "python", "sep": None, "low_memory": False},
                {"encoding": "latin1", "engine": "python", "sep": None, "low_memory": False},
                {"encoding": "utf-16", "engine": "python", "sep": None, "low_memory": False},
                {"encoding": "utf-16-le", "engine": "python", "sep": None, "low_memory": False},
                {"encoding": "utf-16-be", "engine": "python", "sep": None, "low_memory": False},
                {"encoding": "utf-8", "engine": "python", "sep": "\t", "low_memory": False},
                {"encoding": "utf-8", "engine": "python", "sep": None, "compression": "infer", "low_memory": False},
                {"encoding": "utf-8", "engine": "python", "sep": ",", "on_bad_lines": "skip", "dtype": str, "encoding_errors": "replace", "low_memory": False},
            ]:
                try:
                    return pd.read_csv(BytesIO(data), **kwargs)
                except Exception:
                    continue
            # last resort
            return pd.read_csv(BytesIO(data), engine="python", sep=",", on_bad_lines="skip", dtype=str, encoding_errors="replace", low_memory=False)
        if ext == "json":
            return pd.read_json(BytesIO(content))
        if ext == "parquet":
            return pd.read_parquet(BytesIO(content))
        raise HTTPException(status_code=400, detail="Unsupported file format")

    @staticmethod
    def read(
        file: UploadFile,
        sheets: Optional[List[str]] = None,
        merge_identical: bool = False,
    ) -> Dict[str, Any]:
        """Return a structured response for previews: columns, shape, and sample records."""
        ext = DataFactory.detect_extension(file.filename)
        content = DataFactory._to_bytes(file)
        if ext in ("xlsx", "xls"):
            try:
                engine = "xlrd" if ext == "xls" else "openpyxl"
                xls = pd.ExcelFile(BytesIO(content), engine=engine)
            except Exception as e:
                raise HTTPException(status_code=400, detail=str(e))
            if sheets is None or len(sheets) == 0:
                frames = {sn: pd.read_excel(xls, sheet_name=sn, engine=engine) for sn in xls.sheet_names}
            else:
                resolved = []
                for s in sheets:
                    if s in xls.sheet_names:
                        resolved.append(s)
                    else:
                        try:
                            idx = int(s)
                            if 0 <= idx < len(xls.sheet_names):
                                resolved.append(xls.sheet_names[idx])
                        except Exception:
                            pass
                if not resolved:
                    raise HTTPException(status_code=400, detail="No matching sheets found")
                frames = {sn: pd.read_excel(xls, sheet_name=sn, engine=engine) for sn in resolved}
            if merge_identical:
                groups: Dict[tuple, List[str]] = {}
                for name, df in frames.items():
                    key = tuple([c.strip().lower() for c in df.columns])
                    groups.setdefault(key, []).append(name)
                if not groups:
                    raise HTTPException(status_code=400, detail="No sheets to merge")
                largest_key = max(groups.keys(), key=lambda k: len(groups[k]))
                merge_list = groups[largest_key]
                merged = pd.concat([frames[n] for n in merge_list], ignore_index=True)
                return {
                    "merged": True,
                    "merged_sheets": merge_list,
                    "columns": list(merged.columns),
                    "shape": list(merged.shape),
                    "sample": DataResponse.sample_records(merged),
                }
            else:
                same_schema = False
                if len(frames) > 1:
                    keys = [tuple([c.strip().lower() for c in df.columns]) for df in frames.values()]
                    same_schema = all(k == keys[0] for k in keys)
                if same_schema:
                    merged = pd.concat(list(frames.values()), ignore_index=True)
                    return {
                        "merged": True,
                        "merged_sheets": list(frames.keys()),
                        "columns": list(merged.columns),
                        "shape": list(merged.shape),
                        "sample": DataResponse.sample_records(merged),
                    }
                else:
                    detail = {}
                    for name, df in frames.items():
                        detail[name] = {
                            "columns": list(df.columns),
                            "shape": list(df.shape),
                            "sample": DataResponse.sample_records(df),
                        }
                    return {
                        "merged": False,
                        "sheets": detail,
                    }
        if ext == "csv":
            data = content.replace(b"\x00", b"") if b"\x00" in content else content
            for kwargs in [
                {"low_memory": False},
                {"encoding": "utf-8-sig", "engine": "python", "sep": None, "low_memory": False},
                {"encoding": "latin1", "engine": "python", "sep": None, "low_memory": False},
                {"encoding": "utf-16", "engine": "python", "sep": None, "low_memory": False},
                {"encoding": "utf-16-le", "engine": "python", "sep": None, "low_memory": False},
                {"encoding": "utf-16-be", "engine": "python", "sep": None, "low_memory": False},
                {"encoding": "utf-8", "engine": "python", "sep": "\t", "low_memory": False},
                {"encoding": "utf-8", "engine": "python", "sep": None, "compression": "infer", "low_memory": False},
                {"encoding": "utf-8", "engine": "python", "sep": ",", "on_bad_lines": "skip", "dtype": str, "encoding_errors": "replace", "low_memory": False},
            ]:
                try:
                    df = pd.read_csv(BytesIO(data), **kwargs)
                    return DataResponse.single(df)
                except Exception:
                    continue
            df = pd.read_csv(BytesIO(data), engine="python", sep=",", on_bad_lines="skip", dtype=str, encoding_errors="replace", low_memory=False)
            return DataResponse.single(df)
        if ext == "json":
            df = pd.read_json(BytesIO(content))
            return DataResponse.single(df)
        if ext == "parquet":
            df = pd.read_parquet(BytesIO(content))
            return DataResponse.single(df)
        raise HTTPException(status_code=400, detail="Unsupported file format")


class Recommender:
    """Simple heuristics to propose next-step operations based on detected columns."""
    @staticmethod
    def recommend(columns: List[str]) -> List[str]:
        recs = []
        lower_cols = [str(c).lower() for c in columns]
        if any(any(k in c for k in ["revenue", "cost", "price"]) for c in lower_cols):
            recs.append("Financial P&L operations")
        if any("date" in c for c in lower_cols):
            recs.append("Standardize Date")
            recs.append("Sort by Date")
        if any("id" == c or c.endswith("_id") or "account no" in c or "account_no" in c or "account number" in c for c in lower_cols):
            recs.append("Pattern Validation")
        return list(dict.fromkeys(recs))


class DataResponse:
    """Helpers to build preview payloads from DataFrames."""
    @staticmethod
    def sample_records(df: pd.DataFrame, limit: int = 20) -> List[Dict[str, Any]]:
        """Return a limited list of records with NaNs converted to JSON-friendly None."""
        try:
            return df.head(limit).where(pd.notnull(df), None).to_dict(orient="records")
        except Exception:
            return []

    @staticmethod
    def single(df: pd.DataFrame) -> Dict[str, Any]:
        """Wrap a DataFrame into a standard preview structure."""
        return {
            "merged": False,
            "columns": list(df.columns),
            "shape": list(df.shape),
            "sample": DataResponse.sample_records(df),
        }


@app.get("/")
async def root():
    if os.path.isdir(os.path.join(os.path.dirname(__file__), "static")):
        return RedirectResponse(url="/app/")
    return {"status": "ok", "service": "Universal Data Uploader", "endpoints": ["/upload", "/docs"]}


@app.post("/upload")
async def upload(
    file: UploadFile = File(...),
    action: Optional[str] = Form(None),
    sheets: Optional[str] = Form(None),
    merge_identical: Optional[bool] = Form(False),
):
    """
    Upload a file and either list sheets (Excel) or return a preview of columns and sample rows.
    When merge_identical is true, sheets with identical schemas are combined.
    """
    try:
        ext = DataFactory.detect_extension(file.filename)
    except ValueError as e:
        raise HTTPException(status_code=415, detail=str(e))
    if action == "list_sheets":
        if ext not in ("xlsx", "xls"):
            raise HTTPException(status_code=400, detail="Listing sheets is only for Excel files")
        names = DataFactory.list_sheets(file)
        return {"sheets": names}
    selected = None
    if sheets:
        selected = [s.strip() for s in sheets.split(",") if s.strip()]
    try:
        result = DataFactory.read(file, sheets=selected, merge_identical=bool(merge_identical))
    except HTTPException as e:
        # Re-raise FastAPI-aware errors (like unsupported format)
        raise
    except Exception as e:
        # Normalize unexpected parser errors into a client-friendly response
        raise HTTPException(status_code=400, detail=f"Upload failed: {str(e)}")
    if "columns" in result:
        result["recommendations"] = Recommender.recommend(result["columns"])
    elif "sheets" in result:
        recs_per_sheet = {}
        for name, info in result["sheets"].items():
            recs_per_sheet[name] = Recommender.recommend(info.get("columns", []))
        result["recommendations"] = recs_per_sheet
    return result


# -----------------------
# Templates and Pipeline
# -----------------------
TEMPLATES_DIR = os.path.join(os.path.dirname(__file__), "templates")
os.makedirs(TEMPLATES_DIR, exist_ok=True)

# Simple in-memory caches
PREVIEW_CACHE: Dict[str, Dict[str, Any]] = {}
EXPORT_CACHE: Dict[str, Dict[str, Any]] = {}
CACHE_TTL_SECONDS = 15 * 60

REDIS_URL = os.getenv("REDIS_URL")
redis_client = None
if redis_lib and REDIS_URL:
    try:
        redis_client = redis_lib.Redis.from_url(REDIS_URL)
        redis_client.ping()
    except Exception:
        redis_client = None

def _redis_get_bytes(key: str):
    """Fetch cached bytes from Redis by key, if configured."""
    if not redis_client:
        return None
    try:
        return redis_client.get(key)
    except Exception:
        return None

def _redis_set_bytes(key: str, value: bytes, ttl: int = CACHE_TTL_SECONDS):
    """Set cached bytes in Redis with TTL, if configured."""
    if not redis_client:
        return
    try:
        redis_client.setex(key, ttl, value)
    except Exception:
        pass

def _redis_del_prefix(prefix: str):
    """Delete all Redis keys with the given prefix. Returns number of deleted keys."""
    if not redis_client:
        return 0
    deleted = 0
    try:
        for k in redis_client.scan_iter(f"{prefix}*"):
            deleted += int(redis_client.delete(k) or 0)
    except Exception:
        return deleted
    return deleted

def _recipe_fingerprint(steps: List[Dict[str, Any]]) -> str:
    """Build a stable fingerprint for a recipe to power in-memory/Redis caching."""
    try:
        canonical = json.dumps(steps, sort_keys=True, separators=(",", ":"))
    except Exception:
        canonical = str(steps)
    return hashlib.sha256(canonical.encode("utf-8")).hexdigest()

def _cache_get(cache: Dict[str, Dict[str, Any]], key: str):
    """In-memory cache get with TTL enforcement."""
    item = cache.get(key)
    if not item:
        return None
    if (time.time() - item["ts"]) > CACHE_TTL_SECONDS:
        cache.pop(key, None)
        return None
    return item["value"]

def _cache_put(cache: Dict[str, Dict[str, Any]], key: str, value: Any):
    """In-memory cache set with timestamp."""
    cache[key] = {"ts": time.time(), "value": value}
    return value

def _safe_name(name: str) -> str:
    """Normalize a name to a filesystem-safe basename used for template storage."""
    return "".join(ch for ch in name if ch.isalnum() or ch in ("-", "_", " ")).strip().replace(" ", "_")


def apply_recipe(
    df: pd.DataFrame,
    recipe: List[Dict[str, Any]],
    context: Optional[Dict[str, Any]] = None,
) -> Tuple[pd.DataFrame, List[Dict[str, Any]]]:
    """
    Apply a list of steps to a DataFrame using TransformationEngine.
    Returns transformed DataFrame and an audit log.
    """
    eng = TransformationEngine()
    log: List[Dict[str, Any]] = []
    out = df.copy()
    ctx = context or {}
    for idx, step in enumerate(recipe or []):
        stype = step.get("type")
        cfg = step.get("config", {}) or {}
        before = len(out)
        try:
            if stype == "filterDateRange":
                col = cfg.get("col", "Date")
                frm = cfg.get("from")
                to = cfg.get("to")
                tmp = eng.standardize_date(out, [col])
                mask = pd.Series([True] * len(tmp), index=tmp.index)
                if frm:
                    mask &= pd.to_datetime(tmp[col], errors="coerce") >= pd.to_datetime(frm)
                if to:
                    mask &= pd.to_datetime(tmp[col], errors="coerce") <= pd.to_datetime(to)
                out = tmp.loc[mask].reset_index(drop=True)
            elif stype == "filterExpr":
                expr = cfg.get("expr", "")
                out = eng.filter_expr(out, expr=expr)
            elif stype == "joinWithFile":
                key = cfg.get("key")
                how = cfg.get("how", "left")
                suffixes = tuple(cfg.get("suffixes", ["_a", "_b"]))
                join_df = (context or {}).get("join_df")
                if join_df is None:
                    raise ValueError("joinWithFile requires a secondary 'join_file'")
                if not key or key not in out.columns or key not in join_df.columns:
                    raise ValueError("joinWithFile requires 'key' present in both datasets")
                out = out.merge(join_df, on=key, how=how, suffixes=suffixes)
                # no reset index to preserve shape
            elif stype == "mergeColumns":
                cols = cfg.get("cols", [])
                dest = cfg.get("dest", "Merged")
                sep = cfg.get("sep", " ")
                out = eng.merge_columns(out, columns=cols, dest_col=dest, sep=sep, drop_source=False)
            elif stype == "calculateNetProfit":
                out = eng.calculate_net_profit(out)
            elif stype == "standardizeDate":
                col = cfg.get("col", "Date")
                out = eng.standardize_date(out, [col])
            elif stype == "sortByDate":
                col = cfg.get("col", "Date")
                order = cfg.get("order", "asc")
                out = out.sort_values(by=[col], ascending=(order != "desc")).reset_index(drop=True)
            elif stype == "filterDebitCredit":
                col = cfg.get("col", "Type")
                include = cfg.get("include", [])
                out = eng.filter_by_transaction_type(out.rename(columns={col: "Type"}), include=include)
            elif stype == "replaceValues":
                cols = cfg.get("columns", [])
                to_rep = cfg.get("to")
                val = cfg.get("value", "")
                regex = bool(cfg.get("regex", False))
                case = bool(cfg.get("case", True))
                out = eng.replace_values(out, columns=cols, to_replace=to_rep, value=val, regex=regex, case_sensitive=case)
            elif stype == "addColumnFormula":
                dest = cfg.get("dest", "Result")
                expr = cfg.get("expr")
                out = eng.add_column_formula(out, dest_col=dest, expr=expr)
            elif stype == "groupByAggregate":
                by = cfg.get("by", [])
                aggs = cfg.get("aggs", {})
                out = eng.group_by_aggregate(out, by=by, aggs=aggs)
            elif stype == "currencyToFloat":
                cols = cfg.get("columns", [])
                out = eng.currency_to_float(out, columns=cols)
            elif stype == "aggregatePLByPeriod":
                dcol = cfg.get("date_col")
                freq = cfg.get("freq", "M")
                rcol = cfg.get("revenue_col")
                ccol = cfg.get("cogs_col")
                ecol = cfg.get("expenses_col")
                out = eng.aggregate_pl_by_period(out, date_col=dcol, freq=freq, revenue_col=rcol, cogs_col=ccol, expenses_col=ecol)
            elif stype == "dropDuplicates":
                subset = cfg.get("subset")
                keep = cfg.get("keep", "first")
                out = eng.drop_duplicates(out, subset=subset, keep=keep)
            elif stype == "trimWhitespace":
                cols = cfg.get("columns")
                out = eng.trim_whitespace(out, columns=cols)
            elif stype == "lowercaseText":
                cols = cfg.get("columns", [])
                out = eng.lowercase_text(out, columns=cols)
            elif stype == "detectOutliers":
                cols = cfg.get("columns", [])
                k = float(cfg.get("k", 1.5))
                out = eng.detect_outliers_iqr(out, columns=cols, k=k)
            elif stype == "validateRegex":
                column = cfg.get("column")
                pattern = cfg.get("pattern")
                mark_col = cfg.get("mark_col")
                if not column or not pattern:
                    raise ValueError("validateRegex requires 'column' and 'pattern'")
                out = eng.validate_regex(out, column=column, pattern=pattern, mark_col=mark_col)
            else:
                # Unknown step: no-op but capture in log
                pass
            after = len(out)
            log.append({
                "index": idx,
                "type": stype,
                "config": cfg,
                "rows_before": before,
                "rows_after": after,
                "timestamp": datetime.utcnow().isoformat(),
                "dataset_hash": ctx.get("dataset_hash"),
                "template_name": ctx.get("template_name"),
                "py_version": sys.version.split()[0],
                "pandas_version": pd.__version__,
                "platform": platform.platform(),
                "repro_curl": ctx.get("repro_curl"),
            })
        except Exception as ex:
            log.append({
                "index": idx,
                "type": stype,
                "config": cfg,
                "error": str(ex),
                "timestamp": datetime.utcnow().isoformat(),
                "dataset_hash": ctx.get("dataset_hash"),
                "template_name": ctx.get("template_name"),
                "py_version": sys.version.split()[0],
                "pandas_version": pd.__version__,
                "platform": platform.platform(),
                "repro_curl": ctx.get("repro_curl"),
            })
    return out, log


@app.post("/templates/save")
async def save_template(payload: Dict[str, Any] = Body(...)):
    """Persist a named recipe template to disk."""
    name = payload.get("name")
    recipe = payload.get("recipe")
    if not name or not isinstance(recipe, list):
        raise HTTPException(status_code=400, detail="Template requires 'name' and list 'recipe'")
    fname = _safe_name(name)
    path = os.path.join(TEMPLATES_DIR, f"{fname}.json")
    if os.path.exists(path):
        # Overwrite by default for simplicity
        pass
    with open(path, "w", encoding="utf-8") as f:
        json.dump({
            "name": name,
            "schema_version": "1.0.0",
            "recipe": recipe,
            "saved_at": datetime.utcnow().isoformat()
        }, f, ensure_ascii=False, indent=2)
    return {"ok": True, "name": name}


@app.get("/templates")
async def list_templates():
    """List saved template names (without file extensions)."""
    files = [f[:-5] for f in os.listdir(TEMPLATES_DIR) if f.endswith(".json")]
    return {"templates": files}


@app.get("/templates/{name}")
async def get_template(name: str):
    """Load a template by name."""
    fname = _safe_name(name)
    path = os.path.join(TEMPLATES_DIR, f"{fname}.json")
    if not os.path.exists(path):
        raise HTTPException(status_code=404, detail="Template not found")
    with open(path, "r", encoding="utf-8") as f:
        return json.load(f)

@app.post("/templates/migrate")
async def migrate_template(payload: Dict[str, Any] = Body(...)):
    """Migrate a template to a target schema version (no-op for current version)."""
    tpl = payload.get("template")
    target = payload.get("target_version", "1.0.0")
    if not isinstance(tpl, dict):
        raise HTTPException(status_code=400, detail="Requires 'template' object")
    # Current registry only supports 1.0.0 -> 1.0.0 no-op
    tpl["schema_version"] = target
    return {"template": tpl, "changed": False}

@app.post("/templates/rename")
async def rename_template(payload: Dict[str, Any] = Body(...)):
    """Rename a saved template file."""
    old = payload.get("old")
    new = payload.get("new")
    if not old or not new:
        raise HTTPException(status_code=400, detail="Requires 'old' and 'new'")
    oldp = os.path.join(TEMPLATES_DIR, f"{_safe_name(old)}.json")
    newp = os.path.join(TEMPLATES_DIR, f"{_safe_name(new)}.json")
    if not os.path.exists(oldp):
        raise HTTPException(status_code=404, detail="Template not found")
    os.replace(oldp, newp)
    return {"ok": True}

@app.post("/templates/delete")
async def delete_template(payload: Dict[str, Any] = Body(...)):
    """Delete a saved template file."""
    name = payload.get("name")
    if not name:
        raise HTTPException(status_code=400, detail="Requires 'name'")
    p = os.path.join(TEMPLATES_DIR, f"{_safe_name(name)}.json")
    if not os.path.exists(p):
        raise HTTPException(status_code=404, detail="Template not found")
    os.remove(p)
    return {"ok": True}


@app.post("/transform")
async def transform(
    file: UploadFile = File(...),
    recipe: Optional[str] = Form(None),
    template_name: Optional[str] = Form(None),
    sheets: Optional[str] = Form(None),
    merge_identical: Optional[bool] = Form(False),
):
    """
    Execute a full transformation recipe against the uploaded file and return a preview.
    Uses caching for identical dataset+recipe pairs to accelerate subsequent runs.
    """
    steps: List[Dict[str, Any]] = []
    if template_name:
        tpl = await get_template(template_name)
        steps = tpl.get("recipe", [])
    elif recipe:
        try:
            steps = json.loads(recipe)
        except Exception:
            raise HTTPException(status_code=400, detail="Invalid recipe JSON")
    content = DataFactory._to_bytes(file)
    df = DataFactory.read_df(file, sheets=[s.strip() for s in sheets.split(",")] if sheets else None, merge_identical=bool(merge_identical))
    dataset_hash = hashlib.sha256(content).hexdigest()
    fp = _recipe_fingerprint(steps)
    ckey = f"transform:{dataset_hash}:{fp}"
    if redis_client:
        blob = _redis_get_bytes(ckey)
        if blob:
            try:
                return json.loads(blob.decode("utf-8"))
            except Exception:
                pass
    cached = _cache_get(PREVIEW_CACHE, ckey)
    if cached:
        return cached
    repro = "curl -X POST -F \"file=@YOUR_FILE\" -F \"recipe=...\" http://127.0.0.1:8000/transform"
    out, log = apply_recipe(df, steps, context={"dataset_hash": dataset_hash, "template_name": template_name, "repro_curl": repro})
    result = {
        "columns": list(out.columns),
        "shape": list(out.shape),
        "sample": DataResponse.sample_records(out),
        "log": log,
    }
    _cache_put(PREVIEW_CACHE, ckey, result)
    try:
        if redis_client:
            _redis_set_bytes(ckey, json.dumps(result).encode("utf-8"))
    except Exception:
        pass
    return result

@app.post("/preview-transform")
async def preview_transform(
    file: UploadFile = File(...),
    recipe: Optional[str] = Form(None),
    template_name: Optional[str] = Form(None),
    join_file: Optional[UploadFile] = File(None),
    sheets: Optional[str] = Form(None),
    join_sheets: Optional[str] = Form(None),
    limit: Optional[int] = Form(20),
    offset: Optional[int] = Form(0),
):
    """
    Run a transformation in preview mode and return a paginated sample.
    Supports joining with a second file for joinWithFile steps.
    """
    steps: List[Dict[str, Any]] = []
    if template_name:
        tpl = await get_template(template_name)
        steps = tpl.get("recipe", [])
    elif recipe:
        try:
            steps = json.loads(recipe)
        except Exception:
            raise HTTPException(status_code=400, detail="Invalid recipe JSON")
    content = DataFactory._to_bytes(file)
    df = DataFactory.read_df(file, sheets=[s.strip() for s in sheets.split(",")] if sheets else None)
    join_df = None
    join_hash = ""
    if join_file is not None:
        try:
            jbytes = DataFactory._to_bytes(join_file)
            join_hash = hashlib.sha256(jbytes).hexdigest()
            join_df = DataFactory.read_df(join_file, sheets=[s.strip() for s in join_sheets.split(",")] if join_sheets else None)
        except Exception:
            join_df = None
    dataset_hash = hashlib.sha256(content).hexdigest()
    fp = _recipe_fingerprint(steps)
    ckey = f"preview:{dataset_hash}:{join_hash}:{fp}"
    if redis_client:
        rkey = f"{ckey}:{offset}:{limit}"
        blob = _redis_get_bytes(rkey)
        if blob:
            try:
                return json.loads(blob.decode("utf-8"))
            except Exception:
                pass
    cached = _cache_get(PREVIEW_CACHE, ckey)
    if cached:
        out = cached["df"]
        log = cached["log"]
    else:
        out, log = apply_recipe(df, steps, context={"dataset_hash": dataset_hash, "template_name": template_name, "join_df": join_df})
        _cache_put(PREVIEW_CACHE, ckey, {"df": out, "log": log})
    start = max(int(offset or 0), 0)
    end = start + max(int(limit or 20), 1)
    result = {
        "columns": list(out.columns),
        "shape": list(out.shape),
        # Replace NaN with None for JSON safety
        "sample": out.iloc[start:end].where(pd.notnull(out.iloc[start:end]), None).replace({pd.NA: None, float('nan'): None}).to_dict(orient="records"),
        "log": log,
    }
    try:
        if redis_client:
            rkey = f"{ckey}:{offset}:{limit}"
            _redis_set_bytes(rkey, json.dumps(result).encode("utf-8"))
    except Exception:
        pass
    return result

@app.post("/validate")
async def validate(
    file: UploadFile = File(...),
    validators: str = Form(...),
):
    """
    Validate one or more columns against provided regex rules or presets.
    Returns counts and samples of failures per rule.
    """
    presets = {
        "account_no": r"[A-Za-z0-9\-]{6,18}",
        "id_generic": r"[A-Za-z0-9]{6,20}",
        "iban": r"[A-Z]{2}\d{2}[A-Z0-9]{1,30}",
    }
    try:
        rules = json.loads(validators)
    except Exception:
        raise HTTPException(status_code=400, detail="Invalid validators JSON")
    df = DataFactory.read_df(file)
    eng = TransformationEngine()
    results = []
    for rule in rules:
        column = rule.get("column")
        pattern = rule.get("pattern") or presets.get(rule.get("preset",""))
        if not column or not pattern:
            continue
        checked = eng.validate_regex(df, column=column, pattern=pattern)
        mark = [c for c in checked.columns if c.startswith("Valid_") and c.endswith(column)]
        mark = mark[0] if mark else f"Valid_{column}"
        passed = int(checked[mark].sum())
        total = int(checked[mark].count())
        failed = total - passed
        samples = checked.loc[~checked[mark], column].dropna().astype("string").head(10).tolist()
        results.append({"column": column, "pattern": pattern, "passed": passed, "failed": failed, "samples": samples})
    return {"results": results}

@app.post("/profile")
async def profile(
    file: UploadFile = File(...),
    recipe: Optional[str] = Form(None),
    template_name: Optional[str] = Form(None),
    sheets: Optional[str] = Form(None),
    join_file: Optional[UploadFile] = File(None),
    join_sheets: Optional[str] = Form(None),
):
    """
    Generate a lightweight column profile and suggestions.
    Applies an optional recipe first to profile the transformed dataset.
    """
    steps: List[Dict[str, Any]] = []
    if template_name:
        tpl = await get_template(template_name)
        steps = tpl.get("recipe", [])
    elif recipe:
        try:
            steps = json.loads(recipe)
        except Exception:
            raise HTTPException(status_code=400, detail="Invalid recipe JSON")
    content = DataFactory._to_bytes(file)
    df = DataFactory.read_df(file, sheets=[s.strip() for s in sheets.split(",")] if sheets else None)
    join_df = None
    if join_file is not None:
        try:
            join_df = DataFactory.read_df(join_file, sheets=[s.strip() for s in join_sheets.split(",")] if join_sheets else None)
        except Exception:
            join_df = None
    if steps:
        dataset_hash = hashlib.sha256(content).hexdigest()
        out, _ = apply_recipe(df, steps, context={"dataset_hash": dataset_hash, "template_name": template_name, "join_df": join_df})
        df = out
    profile = []
    suggestions = []
    proposed_recipe: List[Dict[str, Any]] = []
    # Sales-oriented heuristics
    cols_lower = [str(c).lower() for c in df.columns]
    # Date candidates
    date_candidates = [c for c in df.columns if any(k in str(c).lower() for k in ["date", "order date", "invoice date", "ship date"])]
    main_date = date_candidates[0] if date_candidates else None
    # Monetary columns
    revenue_col = next((c for c in df.columns if str(c).lower() in ("revenue","sales","sales amount","amount","net sales","total sales")), None)
    cogs_col = next((c for c in df.columns if any(k in str(c).lower() for k in ["cogs","cost of goods sold","cost of goods","cost"])), None)
    expenses_col = next((c for c in df.columns if "expense" in str(c).lower()), None)
    discount_col = next((c for c in df.columns if "discount" in str(c).lower()), None)
    tax_col = next((c for c in df.columns if str(c).lower() in ("tax","vat","gst")), None)
    # Quantity and unit price
    qty_col = next((c for c in df.columns if str(c).lower() in ("qty","quantity","units")), None)
    unit_price_col = next((c for c in df.columns if any(k in str(c).lower() for k in ["unit price","price","rate"])), None)
    for col in df.columns:
        s = df[col]
        dtype = str(s.dtype)
        nulls = int(s.isna().sum())
        uniques = int(s.nunique(dropna=True))
        top = []
        try:
            vc = s.value_counts(dropna=True).head(5)
            top = [{"value": str(i), "count": int(c)} for i, c in vc.items()]
        except Exception:
            pass
        stats = {}
        # Currency and ID pattern heuristics
        str_series = s.astype("string")
        currency_hits = (str_series.str.contains(r"[\$₹€£]|^\(?-?\d{1,3}(?:[ ,]\d{3})+(?:\.\d+)?\)?$", regex=True, na=False)).mean()
        id_hits = (str_series.str.contains(r"^[A-Za-z0-9\-]{6,20}$", regex=True, na=False)).mean()
        if pd.api.types.is_numeric_dtype(s):
            stats = {
                "min": float(pd.to_numeric(s, errors="coerce").min(skipna=True)),
                "max": float(pd.to_numeric(s, errors="coerce").max(skipna=True)),
                "skew": float(pd.to_numeric(s, errors="coerce").skew(skipna=True) or 0),
            }
        elif "date" in col.lower():
            dt = pd.to_datetime(s, errors="coerce")
            stats = {
                "min": dt.min(skipna=True).isoformat() if not dt.isna().all() else None,
                "max": dt.max(skipna=True).isoformat() if not dt.isna().all() else None,
            }
        profile.append({"column": col, "dtype": dtype, "nulls": nulls, "uniques": uniques, "top": top, "stats": stats})
        # Simple health suggestions
        if "date" in col.lower():
            suggestions.append("Standardize Date")
            suggestions.append("Sort by Date")
        if any(k in col.lower() for k in ["revenue", "sales", "turnover", "amount"]):
            suggestions.append("Financial P&L operations")
        if any(k in col.lower() for k in ["cogs", "cost"]):
            suggestions.append("Financial P&L operations")
        if any(k in col.lower() for k in ["id", "account", "iban"]):
            suggestions.append("Pattern Validation")
        # Proposed recipe with confidence
        if currency_hits > 0.5:
            proposed_recipe.append({"type": "calculateNetProfit", "confidence": round(currency_hits, 2)})
        if "date" in col.lower():
            proposed_recipe.append({"type": "standardizeDate", "config": {"col": col}, "confidence": 0.9})
            proposed_recipe.append({"type": "sortByDate", "config": {"col": col}, "confidence": 0.8})
        if id_hits > 0.6:
            proposed_recipe.append({"type": "validateRegex", "config": {"column": col, "pattern": r"^[A-Za-z0-9\\-]{6,20}$"}, "confidence": round(id_hits, 2)})
    # Sales-specific proposed steps
    money_cols = [c for c in [revenue_col, cogs_col, expenses_col, discount_col, tax_col, unit_price_col] if c]
    if money_cols:
        proposed_recipe.append({"type": "currencyToFloat", "config": {"columns": money_cols}, "confidence": 0.9})
    if main_date:
        proposed_recipe.append({"type": "standardizeDate", "config": {"col": main_date}, "confidence": 0.95})
    if revenue_col and cogs_col:
        proposed_recipe.append({"type": "calculateNetProfit", "config": {}, "confidence": 0.95})
    if revenue_col and discount_col:
        # Net Sales = Revenue - Discount (simple)
        proposed_recipe.append({"type": "addColumnFormula", "config": {"dest": "Net Sales", "expr": f"`{revenue_col}` - `{discount_col}`"}, "confidence": 0.85})
    if main_date and (revenue_col or cogs_col or expenses_col):
        proposed_recipe.append({
            "type": "aggregatePLByPeriod",
            "config": {"date_col": main_date, "freq": "M", "revenue_col": revenue_col, "cogs_col": cogs_col, "expenses_col": expenses_col},
            "confidence": 0.9
        })
    # Outlier detection on unit economics
    outlier_targets = [c for c in [unit_price_col, qty_col] if c]
    if outlier_targets:
        proposed_recipe.append({"type": "detectOutliers", "config": {"columns": outlier_targets, "k": 1.5}, "confidence": 0.7})
    # Duplicates
    dup_key = next((c for c in df.columns if any(k in str(c).lower() for k in ["order id","invoice","transaction id","order no","invoice no"])), None)
    if dup_key:
        proposed_recipe.append({"type": "dropDuplicates", "config": {"subset": [dup_key], "keep": "first"}, "confidence": 0.8})
    dataset_hash = hashlib.sha256(content).hexdigest()
    # Optional AI enrichment
    ai_used = False
    ai_provider = None
    try:
        if AI_ENABLED:
            sample_preview = DataResponse.sample_records(df, limit=20)
            ai = _ai_propose_recipe(list(df.columns), sample_preview)
            if ai:
                ai_steps = ai.get("steps", [])
                ai_suggestions = ai.get("suggestions", [])
                ai_provider = ai.get("provider")
                if ai_steps:
                    proposed_recipe = _dedup_steps(proposed_recipe + ai_steps)
                    ai_used = True
                if ai_suggestions:
                    suggestions.extend(ai_suggestions)
    except Exception:
        pass
    # Frequent operations hint
    if steps:
        freq: Dict[str, int] = {}
        for st in steps:
            t = st.get("type")
            if not t:
                continue
            freq[t] = freq.get(t, 0) + 1
        popular = sorted([k for k,v in freq.items() if v >= 2])
        title_map = {
            "standardizeDate": "Standardize Date",
            "sortByDate": "Sort by Date",
            "calculateNetProfit": "Financial P&L operations",
            "validateRegex": "Pattern Validation",
            "addColumnFormula": "Add Column (Formula)",
            "filterExpr": "Filter by Expression",
            "mergeColumns": "Merge Columns",
            "groupByAggregate": "Group & Aggregate",
            "dropDuplicates": "Remove Duplicates",
        }
        for t in popular:
            title = title_map.get(t)
            if title:
                suggestions.append(title)
    return {
        "columns": list(df.columns),
        "shape": list(df.shape),
        "profile": profile,
        "suggestions": list(dict.fromkeys(suggestions + [
            "Monthly P&L",
            "Convert Currency to Number",
            "Detect Outliers",
            "Remove Duplicates",
            "Add Net Sales Column",
        ] if money_cols or main_date else suggestions)),
        "proposed_recipe": proposed_recipe,
        "ai_used": ai_used,
        "ai_provider": ai_provider,
        "dataset_hash": dataset_hash,
    }

# -----------------------
# Join / Entity Resolution Assistant
# -----------------------
@app.post("/join-suggest")
async def join_suggest(
    file_a: UploadFile = File(...),
    file_b: UploadFile = File(...),
    sample: Optional[int] = Form(500),
    sheets_a: Optional[str] = Form(None),
    sheets_b: Optional[str] = Form(None),
):
    """
    Suggest candidate join keys by scoring overlap and uniqueness for common columns.
    Downsamples large datasets for speed.
    """
    sels_a = [s.strip() for s in sheets_a.split(",")] if sheets_a else None
    sels_b = [s.strip() for s in sheets_b.split(",")] if sheets_b else None
    df_a = DataFactory.read_df(file_a, sheets=sels_a)
    df_b = DataFactory.read_df(file_b, sheets=sels_b)
    # Downsample for speed
    if sample and len(df_a) > sample:
        df_a = df_a.sample(sample, random_state=42)
    if sample and len(df_b) > sample:
        df_b = df_b.sample(sample, random_state=42)
    common = list(set(df_a.columns) & set(df_b.columns))
    candidates = []
    for c in common:
        sa = df_a[c]
        sb = df_b[c]
        ua = sa.nunique(dropna=True) / max(len(sa), 1)
        ub = sb.nunique(dropna=True) / max(len(sb), 1)
        # Overlap score: intersection of normalized unique strings
        na = set(sa.dropna().astype("string").str.strip().str.lower().unique().tolist())
        nb = set(sb.dropna().astype("string").str.strip().str.lower().unique().tolist())
        inter = len(na & nb)
        uni = len(na | nb) or 1
        overlap = inter / uni
        # Simple combined score
        score = round(0.5 * overlap + 0.25 * min(ua, ub) + 0.25 * (1 - abs(ua - ub)), 4)
        # Diagnostics for unmatched
        left_only = len(na - nb)
        right_only = len(nb - na)
        candidates.append({
            "column": c,
            "score": score,
            "overlap": round(overlap, 4),
            "left_uniqueness": round(ua, 4),
            "right_uniqueness": round(ub, 4),
            "left_only": left_only,
            "right_only": right_only,
        })
    candidates.sort(key=lambda x: x["score"], reverse=True)
    return {"candidates": candidates}

@app.post("/join-simulate")
async def join_simulate(
    file_a: UploadFile = File(...),
    file_b: UploadFile = File(...),
    key: str = Form(...),
    how: str = Form("left"),
    sample: Optional[int] = Form(10),
    sheets_a: Optional[str] = Form(None),
    sheets_b: Optional[str] = Form(None),
):
    """
    Simulate a join and return diagnostics plus a small sample of the joined rows.
    Uses DuckDB for very large datasets when available.
    """
    sels_a = [s.strip() for s in sheets_a.split(",")] if sheets_a else None
    sels_b = [s.strip() for s in sheets_b.split(",")] if sheets_b else None
    df_a = DataFactory.read_df(file_a, sheets=sels_a)
    df_b = DataFactory.read_df(file_b, sheets=sels_b)
    if key not in df_a.columns or key not in df_b.columns:
        raise HTTPException(status_code=400, detail="Key not present in both files")
    # Convert to duckdb if large and available
    joined = None
    try:
        if duckdb is not None and (len(df_a) > 200000 or len(df_b) > 200000):
            import pyarrow as pa
            con = duckdb.connect()
            con.register("a", pa.Table.from_pandas(df_a, preserve_index=False))
            con.register("b", pa.Table.from_pandas(df_b, preserve_index=False))
            how_sql = {"left":"LEFT", "inner":"INNER", "right":"RIGHT"}.get(how, "LEFT")
            joined = con.execute(f"SELECT * FROM a {how_sql} JOIN b USING({key})").df()
        else:
            joined = df_a.merge(df_b, on=key, how=how, suffixes=("_a","_b"))
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Join failed: {e}")
    # Diagnostics
    left_unmatched = int((~df_a[key].astype("string").isin(df_b[key].astype("string"))).sum())
    right_unmatched = int((~df_b[key].astype("string").isin(df_a[key].astype("string"))).sum())
    cols = list(joined.columns)
    # Ensure JSON-safe sample (replace NaN with None)
    import numpy as np
    head = joined.head(int(sample or 10))
    head = head.where(pd.notnull(head), None)
    head = head.replace({np.nan: None})
    samp = head.to_dict(orient="records")
    return {"columns": cols, "joined_count": int(len(joined)), "left_unmatched": left_unmatched, "right_unmatched": right_unmatched, "sample": samp}

def _apply_high_value_rule(ws, df: pd.DataFrame, value_column: Optional[str], threshold: float):
    if df.empty:
        return
    nrows = df.shape[0]
    ncols = df.shape[1]
    last_letter = get_column_letter(ncols)
    start_row = 2
    end_row = nrows + 1
    fill = PatternFill(start_color="FFF59E", end_color="FFF59E", fill_type="solid")
    if value_column and value_column in df.columns:
        col_idx = df.columns.get_loc(value_column) + 1
        col_letter = get_column_letter(col_idx)
        formula = f"=${col_letter}{start_row}>{threshold}"
        ws.conditional_formatting.add(f"A{start_row}:{last_letter}{end_row}", FormulaRule(formula=[formula], fill=fill))
    else:
        formula = f"=MAX($A{start_row}:${last_letter}{start_row})>{threshold}"
        ws.conditional_formatting.add(f"A{start_row}:{last_letter}{end_row}", FormulaRule(formula=[formula], fill=fill))


@app.post("/export")
async def export_excel(
    file: UploadFile = File(...),
    recipe: Optional[str] = Form(None),
    template_name: Optional[str] = Form(None),
    join_file: Optional[UploadFile] = File(None),
    join_sheets: Optional[str] = Form(None),
    value_column: Optional[str] = Form(None),
    threshold: Optional[float] = Form(10000.0),
    sheets: Optional[str] = Form(None),
    merge_identical: Optional[bool] = Form(False),
    format: Optional[str] = Form("xlsx"),
):
    """
    Export the transformed dataset as Excel/CSV/JSON.
    Adds conditional formatting and a P&L monthly summary sheet when columns exist.
    """
    steps: List[Dict[str, Any]] = []
    if template_name:
        tpl = await get_template(template_name)
        steps = tpl.get("recipe", [])
    elif recipe:
        try:
            steps = json.loads(recipe)
        except Exception:
            raise HTTPException(status_code=400, detail="Invalid recipe JSON")
    content = DataFactory._to_bytes(file)
    df = DataFactory.read_df(file, sheets=[s.strip() for s in sheets.split(",")] if sheets else None, merge_identical=bool(merge_identical))
    join_df = None
    join_hash = ""
    if join_file is not None:
        try:
            jbytes = DataFactory._to_bytes(join_file)
            join_hash = hashlib.sha256(jbytes).hexdigest()
            join_df = DataFactory.read_df(join_file, sheets=[s.strip() for s in join_sheets.split(",")] if join_sheets else None)
        except Exception:
            join_df = None
    dataset_hash = hashlib.sha256(content).hexdigest()
    fp = _recipe_fingerprint(steps)
    ckey = f"export:{dataset_hash}:{join_hash}:{fp}:{value_column}:{threshold}"
    if redis_client:
        blob = _redis_get_bytes(ckey)
        if blob:
            from fastapi.responses import StreamingResponse
            suggested = os.path.splitext(file.filename)[0] + "_transformed.xlsx"
            headers = {"Content-Disposition": f'attachment; filename=\"{suggested}\"'}
            return StreamingResponse(BytesIO(blob), media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers=headers)
    cached = _cache_get(EXPORT_CACHE, ckey)
    if cached:
        from fastapi.responses import StreamingResponse
        suggested = os.path.splitext(file.filename)[0] + "_transformed.xlsx"
        headers = {"Content-Disposition": f'attachment; filename="{suggested}"'}
        return StreamingResponse(BytesIO(cached), media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers=headers)
    repro = "curl -X POST -F \"file=@YOUR_FILE\" -F \"recipe=...\" http://127.0.0.1:8000/export"
    out, log = apply_recipe(df, steps, context={"dataset_hash": dataset_hash, "template_name": template_name, "repro_curl": repro, "join_df": join_df})
    fmt = (format or "xlsx").lower()
    if fmt in ("csv","json"):
        from fastapi.responses import StreamingResponse, PlainTextResponse, Response
        if fmt == "csv":
            text = out.to_csv(index=False)
            headers = {"Content-Disposition": f'attachment; filename="{os.path.splitext(file.filename)[0]}_transformed.csv"'}
            return Response(content=text, media_type="text/csv", headers=headers)
        else:
            text = out.to_json(orient="records")
            headers = {"Content-Disposition": f'attachment; filename="{os.path.splitext(file.filename)[0]}_transformed.json"'}
            return Response(content=text, media_type="application/json", headers=headers)
    # Build P&L Summary (monthly) if columns exist
    rev_col = next((c for c in out.columns if c.lower() in ("revenue","sales","turnover")), None)
    cogs_col = next((c for c in out.columns if c.lower() in ("cogs","cost","cost of goods sold","cost_of_goods_sold")), None)
    exp_col = next((c for c in out.columns if c.lower() == "expenses"), None)
    date_col = next((c for c in out.columns if "date" in c.lower()), None)
    summary = None
    if rev_col and cogs_col and date_col:
        tmp = out.rename(columns={rev_col:"Revenue", cogs_col:"COGS"})
        if exp_col and exp_col in out.columns:
            tmp = tmp.rename(columns={exp_col:"Expenses"})
        try:
            if duckdb is not None and len(tmp) > 200000:
                # DuckDB monthly aggregation
                import pyarrow as pa
                table = pa.Table.from_pandas(tmp, preserve_index=False)
                con = duckdb.connect()
                con.register("t", table)
                has_exp = "Expenses" in tmp.columns
                sql = f"""
                    SELECT date_trunc('month', CAST(Date AS TIMESTAMP)) AS Period,
                           SUM(Revenue) AS Revenue,
                           SUM(COGS) AS COGS
                           {', SUM(Expenses) AS Expenses' if has_exp else ''}
                    FROM (
                        SELECT * FROM t
                    )
                    GROUP BY 1
                    ORDER BY 1
                """
                summary = con.execute(sql).df()
                summary["Gross Profit"] = summary["Revenue"] - summary["COGS"]
                if has_exp:
                    summary["Net Profit"] = summary["Revenue"] - summary["COGS"] - summary["Expenses"]
                else:
                    summary["Net Profit"] = summary["Revenue"] - summary["COGS"]
            else:
                eng = TransformationEngine()
                summary = eng.aggregate_pl_by_period(tmp, date_col=date_col, freq="M", revenue_col="Revenue", cogs_col="COGS", expenses_col="Expenses" if "Expenses" in tmp.columns else None)
        except Exception:
            summary = None
    # Write Excel with conditional formatting, summary, formatting and audit log
    from io import BytesIO
    output = BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        out.to_excel(writer, sheet_name="Data", index=False)
        if summary is not None:
            summary.to_excel(writer, sheet_name="P&L Summary", index=False)
        log_df = pd.DataFrame(log)
        if not log_df.empty:
            log_df.to_excel(writer, sheet_name="Transformation_Log", index=False)
        wb = writer.book
        ws = wb["Data"]
        _apply_high_value_rule(ws, out, value_column=value_column, threshold=float(threshold or 10000))
        # Additional visual cues
        try:
            if "Net Profit" in out.columns:
                idx = out.columns.get_loc("Net Profit")+1
                rng = f"{get_column_letter(idx)}2:{get_column_letter(idx)}{ws.max_row}"
                ws.conditional_formatting.add(rng, ColorScaleRule(start_type='min', start_color='FFF0F9', mid_type='percentile', mid_value=50, mid_color='FFD6E5', end_type='max', end_color='FF1E90FF'))
                # Negative net profit shading
                ws.conditional_formatting.add(rng, FormulaRule(formula=[f"={get_column_letter(idx)}2<0"], fill=PatternFill(start_color='FFFFE5E5', end_color='FFFFE5E5', fill_type='solid')))
            if "Revenue" in out.columns:
                idx = out.columns.get_loc("Revenue")+1
                rng = f"{get_column_letter(idx)}2:{get_column_letter(idx)}{ws.max_row}"
                ws.conditional_formatting.add(rng, DataBarRule(start_type='min', end_type='max', color="FF4DA3FF", showValue="None"))
            # Icon set on Outlier flags if present
            for c in out.columns:
                if c.startswith("Outlier_"):
                    idx = out.columns.get_loc(c)+1
                    rng = f"{get_column_letter(idx)}2:{get_column_letter(idx)}{ws.max_row}"
                    ws.conditional_formatting.add(rng, IconSetRule('3Symbols', 'num', [0,0,0]))
            # Shade weekends in a Date column
            date_like = [c for c in out.columns if "date" in c.lower()]
            if date_like:
                di = out.columns.get_loc(date_like[0]) + 1
                rng = f"A2:{get_column_letter(ws.max_column)}{ws.max_row}"
                # WEEKDAY=1 Sunday; shade if Saturday(7) or Sunday(1)
                for r in range(2, ws.max_row+1):
                    ws.conditional_formatting.add(f"A{r}:{get_column_letter(ws.max_column)}{r}", FormulaRule(formula=[f"=OR(WEEKDAY(${get_column_letter(di)}{r})=1,WEEKDAY(${get_column_letter(di)}{r})=7)"], fill=PatternFill(start_color="FF0F172A", end_color="FF0F172A", fill_type="solid")))
        except Exception:
            pass
        # Formatting policy
        try:
            # Freeze panes
            ws.freeze_panes = "A2"
            # Date formatting and numeric formatting
            date_like = [c for c in out.columns if "date" in c.lower()]
            num_cols = [c for c in out.columns if pd.api.types.is_numeric_dtype(pd.to_numeric(out[c], errors="coerce"))]
            # Locale-aware currency detection from sample strings
            symbol_map = {"$": "$", "€": "€", "£": "£", "₹": "₹"}
            currency_cols = set()
            for c in out.columns:
                samp = pd.Series(out[c].head(30), dtype="string").dropna().astype(str)
                sym_hits = {sym: samp.str.contains(sym).mean() for sym in symbol_map}
                if sym_hits and max(sym_hits.values()) > 0.3:
                    currency_cols.add(c)
            for c in date_like:
                idx = out.columns.get_loc(c) + 1
                col_letter = get_column_letter(idx)
                for row in range(2, ws.max_row + 1):
                    ws[f"{col_letter}{row}"].number_format = "yyyy-mm-dd"
            for c in num_cols:
                idx = out.columns.get_loc(c) + 1
                col_letter = get_column_letter(idx)
                for row in range(2, ws.max_row + 1):
                    if c in currency_cols:
                        ws[f"{col_letter}{row}"].number_format = u'#,##0.00\\ [$]'
                    else:
                        ws[f"{col_letter}{row}"].number_format = "#,##0.00"
            # Approximate autofit
            for i, c in enumerate(out.columns, start=1):
                width = max(len(str(c)), *(len(str(v)) for v in out[c].astype("string").head(100).fillna("").tolist())) + 2
                ws.column_dimensions[get_column_letter(i)].width = min(width, 50)
        except Exception:
            pass
        if "P&L Summary" in wb.sheetnames:
            ws2 = wb["P&L Summary"]
            try:
                ws2.freeze_panes = "A2"
                # Create a simple line chart for Net Profit
                if "Net Profit" in summary.columns:
                    chart = LineChart()
                    chart.title = "Net Profit by Period"
                    chart.y_axis.title = "Net Profit"
                    chart.x_axis.title = "Period"
                    data_ref = Reference(ws2, min_col=summary.columns.get_loc("Net Profit")+1, min_row=1, max_row=summary.shape[0]+1)
                    cat_ref = Reference(ws2, min_col=1, min_row=2, max_row=summary.shape[0]+1)
                    chart.add_data(data_ref, titles_from_data=True)
                    chart.set_categories(cat_ref)
                    ws2.add_chart(chart, "G2")
                # Multi-series chart and Charts sheet
                chart_sheet = wb.create_sheet("Charts")
                mchart = LineChart()
                mchart.title = "Revenue, COGS, Net Profit"
                cols = []
                for name in ["Revenue","COGS","Net Profit"]:
                    if name in summary.columns:
                        cols.append(summary.columns.get_loc(name)+1)
                if cols:
                    for i, col_index in enumerate(cols):
                        data_ref = Reference(ws2, min_col=col_index, min_row=1, max_row=summary.shape[0]+1)
                        mchart.add_data(data_ref, titles_from_data=True)
                    cat_ref = Reference(ws2, min_col=1, min_row=2, max_row=summary.shape[0]+1)
                    mchart.set_categories(cat_ref)
                    chart_sheet.add_chart(mchart, "B2")
                # Hyperlink from Data to P&L Summary
                ws.cell(row=1, column=1, value="P&L Summary").hyperlink = "#'P&L Summary'!A1"
                ws.cell(row=1, column=1).style = "Hyperlink"
            except Exception:
                pass
        if "Transformation_Log" in wb.sheetnames:
            wb["Transformation_Log"].sheet_state = "hidden"
    output.seek(0)
    from fastapi.responses import StreamingResponse
    suggested = os.path.splitext(file.filename)[0] + "_transformed.xlsx"
    headers = {"Content-Disposition": f'attachment; filename="{suggested}"'}
    # Put into cache
    _cache_put(EXPORT_CACHE, ckey, output.getvalue())
    try:
        if redis_client:
            _redis_set_bytes(ckey, output.getvalue())
    except Exception:
        pass
    return StreamingResponse(output, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers=headers)

@app.post("/cache/clear")
async def cache_clear(prefix: Optional[str] = Form(None)):
    """Clear in-memory and Redis caches. Optionally limit by key prefix."""
    PREVIEW_CACHE.clear()
    EXPORT_CACHE.clear()
    cleared = 0
    if redis_client:
        if prefix:
            cleared += _redis_del_prefix(prefix)
        else:
            for p in ["export:", "preview:", "transform:"]:
                cleared += _redis_del_prefix(p)
    return {"ok": True, "redis_deleted": cleared}

if __name__ == "__main__":
    uvicorn.run(app, host="127.0.0.1", port=8000)

